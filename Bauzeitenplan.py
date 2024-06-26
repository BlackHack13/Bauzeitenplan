import tkinter
import tkinter.messagebox
import tkinter.filedialog
import customtkinter
from PIL import Image
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import keyboard
import calendar
import holidays
from datetime import date, timedelta
import requests


PATH = os.path.dirname(os.path.realpath(__file__)) + "/IMG"
customtkinter.set_appearance_mode("System")  # Modes: "System", "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue", "green", "dark-blue"
MONATE = ("Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember")
VERSION = 6.0

class GUI(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        check_version_window = CheckVersion(self)
        check_version_window.check_for_new_version()
        if check_version_window.winfo_exists():
            check_version_window.lift()
            self.wait_window(check_version_window)

        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        aufloesung_faktor = {
            (1280, 720): (7, 1000, 800),  # HD
            (1366, 768): (7, 1000, 800),
            (1440, 900): (7, 1000, 800),
            (1600, 900): (7, 1000, 800),
            (1920, 1080): (6, 900, 800),  # FullHD
            (1680, 1050): (6, 900, 800),
            (1400, 1050): (6, 900, 800),
            (2048, 1152): (6, 900, 800),
            (2560, 1440): (5.5, 800, 800),  # QHD
            (2560, 1600): (5.5, 800, 800),
            (3200, 1800): (5, 800, 800),
            (3840, 2160): (5, 800, 800)}  # UHD
        faktor_und_groesse = aufloesung_faktor.get((screen_width, screen_height), (6, 900, 800))
        WIDTH = faktor_und_groesse[1]
        HEIGHT = faktor_und_groesse[2]

        self.title("Bauzeitenplan Creator")
        self.iconbitmap(PATH + '/icon.ico')
        self.geometry(f"{WIDTH}x{HEIGHT}")
        self.windows = {}
        self.protocol("WM_DELETE_WINDOW", self.on_main_window_close)

        # create the frames
        self.frame_oben = customtkinter.CTkFrame(self)
        self.frame_links = customtkinter.CTkFrame(self, corner_radius=0)
        self.frame_mitte = customtkinter.CTkFrame(self)
        self.frame_rechts = customtkinter.CTkFrame(self)

        # grid the frames
        self.frame_oben.grid(row=0, column=0, columnspan=3, sticky="nsew")
        self.frame_links.grid(row=1, column=0, padx=(0, 20), sticky="nsew")
        self.frame_mitte.grid(row=1, column=1, padx=(0, 20), pady=(10, 0), sticky="nsew")
        self.frame_rechts.grid(row=1, column=2, pady=(10, 0), sticky="nsew")

        # ==== grid layout (1x1) oben ===
        self.frame_oben.grid_rowconfigure(1, weight=1)
        self.frame_oben.grid_columnconfigure(1, weight=1)

        # configure grid layout (2x3)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(3, weight=0)
        self.grid_rowconfigure(1, weight=1)

        # ==== grid layout (1x11) L ===
        self.frame_links.grid_rowconfigure(0, minsize=10)
        self.frame_links.grid_rowconfigure(5, weight=1)
        self.frame_links.grid_rowconfigure(8, minsize=20)
        self.frame_links.grid_rowconfigure(20, minsize=10)

        # === grid layout (3x7) M ===
        self.frame_mitte.rowconfigure((1, 2, 3), weight=1)
        self.frame_mitte.rowconfigure(8, weight=10)
        self.frame_mitte.columnconfigure((0, 1), weight=1)
        self.frame_mitte.columnconfigure(2, weight=0)

        # === grid layout (1x11) R ===
        self.frame_rechts.rowconfigure(0, weight=1)
        self.frame_rechts.rowconfigure(8, weight=10)
        self.frame_rechts.columnconfigure(8, weight=1)
        self.frame_rechts.columnconfigure(2, weight=0)

        # ////// Top
        self.logo = customtkinter.CTkImage(light_image=Image.open(PATH + "/logo.png"), dark_image=Image.open(PATH + "/logo.png"), size=(140, 60))
        self.label_logo = customtkinter.CTkLabel(master=self.frame_oben, image=self.logo, textvariable="", text="")
        self.label_logo.grid(row=0, column=0, padx=(5, 0), pady=(5, 0))

        # ////// Links
        self.label_darstellung = customtkinter.CTkLabel(self.frame_links, text="Darstellung:", anchor="w")
        self.label_darstellung.grid(row=6, column=0)
        self.darstellung_menu = customtkinter.CTkOptionMenu(self.frame_links, values=["Light", "Dark", "System"], command=self.change_appearance_mode)
        self.darstellung_menu.grid(row=7, column=0, pady=(0, 20))
        self.darstellung_menu.set("System")

        self.scaling_label = customtkinter.CTkLabel(self.frame_links, text="UI Skalierung:", anchor="w")
        self.scaling_label.grid(row=8, column=0)
        self.scaling_menu = customtkinter.CTkOptionMenu(self.frame_links, values=["50%", "75%", "100%", "125%", "150%", "200%"], command=self.change_scaling)
        self.scaling_menu.grid(row=9, column=0, padx=20, pady=(0, 20))
        self.scaling_menu.set("100%")

        # ////// Mitte
        self.bauherr_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Bauherr: ", width=300, height=25, border_width=2, corner_radius=5)
        self.bauherr_entry.grid(row=1, column=0, columnspan=3, padx=(5, 5))
        self.projektname_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Projektname: ", width=300, height=25, border_width=2, corner_radius=5)
        self.projektname_entry.grid(row=2, column=0, columnspan=3, padx=(5, 5))
        self.projektnummer_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Projektnummer: ", width=300, height=25, border_width=2, corner_radius=5)
        self.projektnummer_entry.grid(row=3, column=0, columnspan=3, padx=(5, 5))

        self.startjahr_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Startjahr: ", width=120, height=25, border_width=2, corner_radius=5)
        self.startjahr_entry.grid(row=5, column=0, padx=(5, 5), pady=15)
        self.endjahr_entry = customtkinter.CTkEntry(master=self.frame_mitte, placeholder_text="Endjahr: ", width=120, height=25, border_width=2, corner_radius=5)
        self.endjahr_entry.grid(row=5, column=1, padx=(5, 5), pady=15)

        self.startmonat_entry = customtkinter.CTkOptionMenu(master=self.frame_mitte, values=MONATE)
        self.startmonat_entry.grid(row=6, column=0, padx=(5, 1))
        self.endmonat_entry = customtkinter.CTkOptionMenu(master=self.frame_mitte, values=MONATE)
        self.endmonat_entry.grid(row=6, column=1, padx=(1, 5))

        self.wochentage = customtkinter.CTkCheckBox(master=self.frame_mitte, text="Wochentage")
        self.wochentage.grid(row=7, column=0, pady=10)

        self.button_speichern = customtkinter.CTkButton(master=self.frame_mitte, text="Speichern", width=300, height=35, corner_radius=5, command=self.save_file)
        self.button_speichern.grid(row=9, column=0, padx=(5, 1), pady=(0, 60))
        self.button_open = customtkinter.CTkButton(master=self.frame_mitte, text="Öffnen in Excel", width=300, height=35, corner_radius=5, command=self.open, state="disabled")
        self.button_open.grid(row=9, column=1, padx=(1, 5), pady=(0, 60))

        self.entry_filepath = tkinter.Entry(self.frame_mitte, text="", state="readonly", width=100)
        self.entry_filepath.grid(row=10, column=0, columnspan=2, padx=(5, 1), pady=(0, 20))
        self.entry_filepath.grid_remove()

        # ////// Rechts
        self.txt_without_weekdays = customtkinter.CTkLabel(master=self.frame_rechts, text="Ohne Wochentage")
        self.txt_without_weekdays.grid(row=1, column=0, padx=(5, 5))
        self.img_without_weekdays = customtkinter.CTkImage(light_image=Image.open(PATH + "/prev1.png"), dark_image=Image.open(PATH + "/prev1.png"), size=(320, 170))
        self.label_img_without_weekdays = customtkinter.CTkLabel(master=self.frame_rechts, image=self.img_without_weekdays, text="")
        self.label_img_without_weekdays.grid(row=2, column=0, padx=(5, 5), pady=(0, 20))

        self.txt_weekdays = customtkinter.CTkLabel(master=self.frame_rechts, text="Mit Wochentage")
        self.txt_weekdays.grid(row=6, column=0, padx=(5, 5))
        self.img_baukosten = customtkinter.CTkImage(light_image=Image.open(PATH + "/prev2.png"), dark_image=Image.open(PATH + "/prev2.png"), size=(320, 170))
        self.label_img_weekdays = customtkinter.CTkLabel(master=self.frame_rechts, image=self.img_baukosten, text="")
        self.label_img_weekdays.grid(row=7, column=0, padx=(5, 5))

        self.label_name = customtkinter.CTkLabel(master=self.frame_rechts, text="J. Winkler", anchor="n").grid(row=9, column=0)
        self.label_version = customtkinter.CTkLabel(master=self.frame_rechts, text=("Version: " + str(VERSION)), anchor="n").grid(row=10, column=0, pady=(0, 15))

        self.open_excel_file = False

    def open(self):
        if self.open_excel_file:
            os.startfile(self.filepath)

    def save_file(self):
        eingabe = self.get_current_entries()
        if eingabe is None:
            return

        dateiname = eingabe.projektnummer + " Bauzeitenplan"
        self.filepath = tkinter.filedialog.asksaveasfilename(
            initialfile=dateiname,
            defaultextension=".xlsx",
            filetypes=[("Excel Arbeitsmappe", ".xlsx")])
        if self.filepath:
            workbook = Excel(eingabe).get_workbook()
            workbook.save(self.filepath)
            self.open_excel_file = True

            self.button_open.configure(state="normal")
            self.entry_filepath.configure(state="normal")
            self.entry_filepath.delete(0, tkinter.END)
            filepath_without_ext = os.path.splitext(self.filepath)[0]
            self.entry_filepath.insert(0, filepath_without_ext)
            self.entry_filepath.configure(state="readonly")
            self.entry_filepath.grid()

    def get_current_entries(self):
        eingabe = Eingabe()
        eingabe.bauherr = self.bauherr_entry.get()
        eingabe.projektname = self.projektname_entry.get()
        eingabe.projektnummer = self.projektnummer_entry.get()
        eingabe.wochentage = bool(self.wochentage.get())
        number1 = self.startjahr_entry.get()
        number2 = self.endjahr_entry.get()

        if not number1.isdigit() or not number2.isdigit():
            tkinter.messagebox.showerror("Fehler", "Jahre müssen eine Zahl sein")
            return None
        if int(number1) < 2000 or int(number2) < 2000:
            tkinter.messagebox.showerror("Fehler", "Jahre müssen ab 2000 sein")
            return None
        if int(number2) < int(number1):
            tkinter.messagebox.showerror("Fehler", "Startjahr muss kleiner oder gleich dem Endjahr sein.")
            return None
        eingabe.startjahr = int(number1)
        eingabe.endjahr = int(number2)

        eingabe.startmonat = MONATE.index(self.startmonat_entry.get()) + 1
        eingabe.endmonat = MONATE.index(self.endmonat_entry.get()) + 1
        if eingabe.startjahr == eingabe.endjahr and eingabe.startmonat > eingabe.endmonat:
            tkinter.messagebox.showerror("Fehler", "Startmonat muss kleiner Endmonat sein, wenn Start- und Endjahr gleich sind.")
            return None

        return eingabe

    # GUI Funktionen
    @staticmethod
    def change_appearance_mode(new_appearance_mode):
        customtkinter.set_appearance_mode(new_appearance_mode)

    @staticmethod
    def change_scaling(new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)

    def on_main_window_close(self):
        self.destroy()

    def toggle_fullscreen(self, event=None):
        self.attributes('-fullscreen', not self.attributes('-fullscreen'))
        if self.attributes('-fullscreen'):
            self.state('zoomed')  # Fenster maximieren
            self.wm_attributes('-topmost', 1)  # Fenster immer im Vordergrund anzeigen
        else:
            self.state('normal')  # Fensterzustand auf normal zurücksetzen
            self.wm_attributes('-topmost', 0)  # Standardmäßige Verhaltensweise wiederherstellen

class CheckVersion(customtkinter.CTkToplevel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.geometry("430x60")
        self.title("Neue Version")
        self.iconbitmap(PATH + '/icon.ico')
        self.current_version = float(VERSION)
    def check_for_new_version(self):
        latest_version = self.get_latest_version_from_github()
        if latest_version is None:
            customtkinter.CTkLabel(self, text="Fehler beim Abrufen des Versionsabgleiches.").pack()
            customtkinter.CTkButton(self, text="OK", command=self.destroy).pack()
            self.lift()
            return

        if latest_version > self.current_version:
            customtkinter.CTkLabel(self, text=f"Eine neue Version ({latest_version}) ist verfügbar! Bitte aktualisiere die Anwendung.").pack()
            customtkinter.CTkButton(self, text="OK", command=self.destroy).pack()
            self.lift()
        else:
            self.destroy()

    def get_latest_version_from_github(self):
        try:
            response = requests.get("https://api.github.com/repos/BlackHack13/Bauzeitenplan/releases/latest")
            response.raise_for_status()
            data = response.json()
            latest_version = float(data["tag_name"].lstrip('v'))
            return latest_version
        except Exception:
            return None


    def get_latest_version_from_github(self):
        try:
            response = requests.get("https://api.github.com/repos/BlackHack13/Bauzeitenplan/releases/latest")
            response.raise_for_status()
            data = response.json()
            latest_version = float(data["tag_name"].lstrip('v'))
            return latest_version
        except Exception:
            return None


class Eingabe:
    def __init__(self):
        self.bauherr = ""
        self.projektname = ""
        self.projektnummer = ""
        self.startjahr = -1
        self.endjahr = -1
        self.startmonat = -1
        self.endmonat = -1
        self.wochentage = False

class Excel:
    def __init__(self, eingabe):
        self.START_VARIABLER_ANHANG = 7
        self.START_BESCHRIFTUNG_VARIABLE = openpyxl.utils.get_column_letter(self.START_VARIABLER_ANHANG - 1)

        self.workbook = openpyxl.Workbook()
        self.eingabe = eingabe
        worksheet = self.workbook.active
        worksheet.title = eingabe.projektnummer + " Bauzeitenplan"

        # === Basics===
        worksheet['A1'] = "Projektnummer:"
        worksheet['B1'] = eingabe.projektnummer
        worksheet['A2'] = "Projektname:"
        worksheet['B2'] = eingabe.projektname
        worksheet['A3'] = "Bauherr:"
        worksheet['B3'] = eingabe.bauherr
        worksheet['A4'] = "Bauzeit: " + str(len(set((x[0], x[1]) for x in self._create_calender()))) + " Monate"
        worksheet['A5'] = "Stand: " + date.today().strftime("%d.%m.%Y")

        if eingabe.wochentage:
            row = 6
        else:
            row = 5
        labels = [("C", "Startdatum:"), ("D", "Enddatum:"), ("E", "Gesamter Zeitraum:")]
        for col, label in labels:
            worksheet[f"{col}{row}"] = label


        # === Beschriftung Calender===
        worksheet[self.START_BESCHRIFTUNG_VARIABLE + '3'] = "Jahr"
        worksheet[self.START_BESCHRIFTUNG_VARIABLE + '4'] = "Monat"
        worksheet[self.START_BESCHRIFTUNG_VARIABLE + '5'] = "Kalenderwoche"
        if eingabe.wochentage:
            worksheet['A7'] = "Anforderungen:"
            worksheet[self.START_BESCHRIFTUNG_VARIABLE + '6'] = "Wochentag"
            worksheet[self.START_BESCHRIFTUNG_VARIABLE + '7'] = "Datum"
        else:
            worksheet['A6'] = "Anforderungen:"



        self.calender_to_excel(self.START_VARIABLER_ANHANG)
        self.format_excel(self.START_BESCHRIFTUNG_VARIABLE, self.START_VARIABLER_ANHANG)

    def get_workbook(self):
        return self.workbook

    def _create_calender(self):
        wochentage = ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag']

        start_date = date(self.eingabe.startjahr, self.eingabe.startmonat, 1)
        end_date = date(self.eingabe.endjahr, self.eingabe.endmonat, 1)
        end_date = end_date.replace(day=calendar.monthrange(end_date.year, end_date.month)[1])

        current_date = start_date
        calendar_list = []

        while current_date <= end_date:
            year = current_date.year
            month = current_date.month
            week_number = current_date.isocalendar()[1]
            if self.eingabe.wochentage:
                weekday = wochentage[current_date.weekday()]
                calendar_list.append((year, month, week_number, weekday, current_date.strftime("%d"), current_date))
            else:
                calendar_list.append((year, month, week_number))
            current_date += timedelta(days=1)

        return calendar_list

    @staticmethod
    def _format_month(month):
        return f"{MONATE[month - 1]}"

    def calender_to_excel(self, START_VARIABLER_ANHANG):
        worksheet = self.workbook.worksheets[0]
        calendar_days = self._create_calender()
        spaltennummer_start = START_VARIABLER_ANHANG
        spaltennummer = spaltennummer_start - 1

        last_week_number = None
        de_holidays = holidays.Germany(state='BW') # States: BB, BE, BW, BY, BYP, HB, HE, HH, MV, NI, NW, RP, SH, SL, SN, ST, TH

        for day_info in calendar_days:
            year = day_info[0]
            month = day_info[1]
            week_number = day_info[2]

            if not self.eingabe.wochentage and week_number == last_week_number:
                continue

            spaltennummer += 1
            last_week_number = week_number

            worksheet.cell(row=3, column=spaltennummer).value = year
            worksheet.cell(row=4, column=spaltennummer).value = self._format_month(month)
            worksheet.cell(row=5, column=spaltennummer).value = "KW " + str(week_number)

            if self.eingabe.wochentage:
                weekday = day_info[3]
                day = day_info[4]
                current_date = day_info[5]
                worksheet.cell(row=6, column=spaltennummer).value = weekday
                worksheet.cell(row=7, column=spaltennummer).value = day

                holiday_name = de_holidays.get(current_date)
                if holiday_name:
                    worksheet.cell(row=8, column=spaltennummer).value = holiday_name

    def format_excel(self, START_BESCHRIFTUNG_VARIABLE, START_VARIABLER_ANHANG):
        Format_Excel(self.eingabe, self.workbook, START_BESCHRIFTUNG_VARIABLE, START_VARIABLER_ANHANG).run_format_excel()

from openpyxl.styles import PatternFill

class Format_Excel:
    LAENGE_SPALTE = 30

    def __init__(self, eingabe, workbook, START_BESCHRIFTUNG_VARIABLE, START_VARIABLER_ANHANG):
        self.eingabe = eingabe
        self.workbook = workbook
        self.START_BESCHRIFTUNG_VARIABLE_STR = START_BESCHRIFTUNG_VARIABLE
        self.START_VARIABLER_ANHANG_INT = START_VARIABLER_ANHANG
        self.START_VARIABLER_ANHANG_STR = openpyxl.utils.get_column_letter(self.START_VARIABLER_ANHANG_INT)

    def _merge_left_layoutcell(self, worksheet):
        initial_cells_to_merge = ['B1:E1', 'B2:E2', 'B3:E3']
        for cell_range in initial_cells_to_merge:
            worksheet.merge_cells(cell_range)

        if self.eingabe.wochentage:
            cells_to_merge = ['A7:B7', 'B4:E5', 'A6:B6', 'C6:C7', 'D6:D7', 'E6:E7']
        else:
            cells_to_merge = ['A6:B6', 'B4:B5', 'C5:C6', 'D5:D6', 'E5:E6', 'C4:E4']
        for cell_range in cells_to_merge:
            worksheet.merge_cells(cell_range)

        self._cell_alignment(worksheet)

    def _cell_alignment(self, worksheet):
        row = 6 if self.eingabe.wochentage else 5
        cells_to_format = [f"{col}{row}" for col in ["C", "D", "E"]]
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in cells_to_format:
            worksheet[cell].alignment = alignment

        row = 7 if self.eingabe.wochentage else 6
        worksheet[f"A{row}"].alignment = alignment

    def _column_width(self, worksheet):
        columns = { "A": 17, "B": 20, "C": 11, "D": 10, "E": 11, self.START_BESCHRIFTUNG_VARIABLE_STR: 14}
        for col_num, col_width in columns.items():
            worksheet.column_dimensions[col_num].width = col_width

    def _count_up_left_layoutcell(self, worksheet):
        durchnum = 0
        start_nummerrierung = 8 if self.eingabe.wochentage else 7

        for row in range(start_nummerrierung, Format_Excel.LAENGE_SPALTE):
            durchnum += 1
            cell = f"A{row}"
            worksheet.merge_cells(f"{cell}:B{row}")
            worksheet[cell] = f"{durchnum}."

    def _title_layoutcell(self, worksheet):
        worksheet[self.START_VARIABLER_ANHANG_STR + '1'] = "Bauzeitenplan"
        worksheet.merge_cells(start_row=1, start_column=self.START_VARIABLER_ANHANG_INT, end_row=2, end_column=worksheet.max_column)
        worksheet[self.START_VARIABLER_ANHANG_STR + '1'].alignment = Alignment(horizontal='center', vertical='center')

    def _font_size(self, worksheet):
        header1 = Font(bold=True)
        if self.eingabe.wochentage:
            worksheet['A7'].font = header1
            worksheet[self.START_BESCHRIFTUNG_VARIABLE_STR + '6'].font = header1
            worksheet[self.START_BESCHRIFTUNG_VARIABLE_STR + '7'].font = header1
        else:
            worksheet['A6'].font = header1

        worksheet[self.START_BESCHRIFTUNG_VARIABLE_STR + '3'].font = header1
        worksheet[self.START_BESCHRIFTUNG_VARIABLE_STR + '4'].font = header1
        worksheet[self.START_BESCHRIFTUNG_VARIABLE_STR + '5'].font = header1

        header2 = Font(bold=True, size=18)
        worksheet[self.START_VARIABLER_ANHANG_STR + '1'].font = header2

    def _merge_and_center_cells_in_calender(self, worksheet, row_num):
        start_col = None
        current_value = None

        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_num, column=col)
            if current_value is None or cell.value != current_value:
                if start_col is not None and start_col != col - 1:
                    worksheet.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=col - 1)
                    merged_cell = worksheet.cell(row=row_num, column=start_col)
                    merged_cell.alignment = Alignment(horizontal='center', vertical='center')

                start_col = col
                current_value = cell.value

        if start_col is not None and start_col != worksheet.max_column:
            worksheet.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=worksheet.max_column)
            merged_cell = worksheet.cell(row=row_num, column=start_col)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')

    def _color_weekends(self, worksheet):
        light_orange_fill = PatternFill(start_color="FFFFFF", end_color="FFFACD", fill_type="lightVertical")
        for col in range(self.START_VARIABLER_ANHANG_INT, worksheet.max_column + 1):
            weekday_cell = worksheet.cell(row=6, column=col)
            if weekday_cell.value in ["Samstag", "Sonntag"]:
                for row in range(6, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=col).fill = light_orange_fill

    def _add_borders_to_all_cells(self, worksheet):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    def _add_thick_border_new_year(self, worksheet):
        thick_border = Border(left=Side(style='thin'),
                              right=Side(style='thick'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        previous_year = None
        for col in range(self.START_VARIABLER_ANHANG_INT, worksheet.max_column):
            year_cell = worksheet.cell(row=3, column=col)
            next_year_cell = worksheet.cell(row=3, column=col + 1)

            if year_cell.value != previous_year and next_year_cell.value is not None and year_cell.value is not None and next_year_cell.value != year_cell.value:
                for row in range(3, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=col).border = thick_border
                previous_year = year_cell.value
    def run_format_excel(self):
        worksheet = self.workbook.worksheets[0]
        self._merge_left_layoutcell(worksheet)
        self._count_up_left_layoutcell(worksheet)
        self._title_layoutcell(worksheet)
        self._font_size(worksheet)
        self._column_width(worksheet)
        self._add_borders_to_all_cells(worksheet)
        self._add_thick_border_new_year(worksheet)
        self._color_weekends(worksheet)

        for row in [3, 4, 5]:
            self._merge_and_center_cells_in_calender(worksheet, row)



class Keyboard_Input:
    def __init__(self, gui):
        self.gui = gui

    def keyboard_input(self, event):
        if event.event_type == keyboard.KEY_DOWN:
            if event.name == 'f11':                                 # F11 Vollbild
                self.gui.toggle_fullscreen()

            elif keyboard.is_pressed('ctrl') and event.name == 'q':  # Strg+ Q Fenster schließen
                self.gui.on_main_window_close()

            elif keyboard.is_pressed('ctrl') and event.name == 'o':  # Strg+ O Datei öffnen
                self.gui.open()

            elif keyboard.is_pressed('ctrl') and event.name == 's':  # Strg+ S Datei speichern
                self.gui.save_file()

    def start_keyboard_input(self):
        keyboard.on_press(self.keyboard_input)

def main():
    gui = GUI()
    keyboard_input = Keyboard_Input(gui)
    keyboard_input.start_keyboard_input()
    gui.mainloop()

if __name__ == "__main__":
    main()
